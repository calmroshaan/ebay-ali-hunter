# output/exporter.py
# -----------------------------------------------
# Saves results to Excel with 2 sheets:
#   Sheet 1 — Profitable Matches (eBay + AliExpress)
#   Sheet 2 — eBay Top Sellers (market research)
# -----------------------------------------------

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR
from utils.logger import get_logger

logger = get_logger(__name__)

# --- Colors ---
COLOR_GREEN       = "C6EFCE"
COLOR_YELLOW      = "FFEB9C"
COLOR_BLUE        = "BDD7EE"
COLOR_HEADER_BG   = "2F4F7F"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_EBAY_HEADER = "1B4F72"

# --- Sheet 1 columns: Profitable Matches ---
MATCH_COLUMNS = [
    ("Keyword",         "keyword",       20),
    ("Market",          "market",        10),
    ("Currency",        "currency",      10),
    ("eBay Title",      "title",         40),
    ("eBay Price",      "ebay_price",    12),
    ("eBay Shipping",   "ebay_shipping", 14),
    ("eBay Fee",        "ebay_fee",      12),
    ("eBay URL",        "ebay_url",      30),
    ("Ali Title",       "ali_title",     40),
    ("Ali Price",       "ali_price",     12),
    ("Ali Shipping",    "ali_shipping",  14),
    ("Ali URL",         "ali_url",       30),
    ("Profit",          "profit",        12),
    ("Margin %",        "margin_pct",    12),
    ("Match Score",     "match_score",   12),
    ("Sold Count",      "sold_count",    12),
    ("Seller Rating",   "seller_rating", 14),
    ("Welcome Deal",    "welcome_deal",  14),
]

# --- Sheet 2 columns: eBay Top Sellers ---
EBAY_COLUMNS = [
    ("Keyword",         "keyword",       20),
    ("Market",          "market",        10),
    ("eBay Title",      "title",         40),
    ("eBay Price",      "ebay_price",    12),
    ("eBay Shipping",   "ebay_shipping", 14),
    ("Sold Count",      "sold_count",    12),
    ("Seller Rating",   "seller_rating", 14),
    ("Welcome Deal",    "welcome_deal",  14),
    ("eBay URL",        "ebay_url",      30),
]


def export_results(results: list[dict], ebay_top: list[dict] = None) -> str | None:
    """
    Takes profitable matches and eBay top sellers.
    Writes both to separate sheets in one Excel file.
    Returns file path or None if nothing to save.
    """
    if not results and not ebay_top:
        logger.info("No results to export")
        return None

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename  = f"results_{timestamp}.xlsx"
    filepath  = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()

    # --- Sheet 1: Profitable Matches ---
    ws1 = wb.active
    ws1.title = "Profitable Matches"
    _write_sheet(ws1, results or [], MATCH_COLUMNS, is_match_sheet=True)

    # --- Sheet 2: eBay Top Sellers ---
    ws2 = wb.create_sheet(title="eBay Top Sellers")
    _write_sheet(ws2, ebay_top or [], EBAY_COLUMNS, is_match_sheet=False)

    try:
        wb.save(filepath)
        logger.info(f"Exported {len(results or [])} matches + {len(ebay_top or [])} eBay top sellers -> {filepath}")
        return filepath
    except Exception as e:
        logger.error(f"Export failed: {e}", exc_info=True)
        return None


def _write_sheet(ws, data: list[dict], columns: list, is_match_sheet: bool):
    """Write header and data rows to a worksheet."""
    _write_header(ws, columns)

    for row_idx, product in enumerate(data, start=2):
        _write_row(ws, row_idx, product, columns, is_match_sheet)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_header(ws, columns: list):
    """Write styled header row."""
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    thin_border = _get_border()

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20


def _write_row(ws, row_idx: int, product: dict, columns: list, is_match_sheet: bool):
    """Write one product row with color coding."""
    margin  = product.get("margin_pct", 0)
    welcome = product.get("welcome_deal", False)

    # Color logic
    if is_match_sheet:
        if welcome:
            fill = PatternFill("solid", fgColor=COLOR_BLUE)
        elif margin >= 40:
            fill = PatternFill("solid", fgColor=COLOR_GREEN)
        else:
            fill = PatternFill("solid", fgColor=COLOR_YELLOW)
    else:
        # eBay Top Sellers sheet — blue for welcome deal, white otherwise
        if welcome:
            fill = PatternFill("solid", fgColor=COLOR_BLUE)
        else:
            fill = PatternFill("solid", fgColor="F2F2F2")

    thin_border = _get_border()

    for col_idx, (_, key, _) in enumerate(columns, start=1):
        value = product.get(key, "")

        if isinstance(value, float):
            value = round(value, 2)

        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Clickable URLs
        if key in ("ebay_url", "ali_url") and value:
            cell.hyperlink = str(value)
            cell.font      = Font(color="0000FF", underline="single")


def _get_border():
    """Thin border for all cells."""
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)